#!/usr/bin/env python
# coding: utf-8
import pandas as pd
import numpy as np
#import pathlib
#import datetime
#import xlrd
import openpyxl
import configparser
import os
import copy

class rschedule:
    'Class for Revit schedules'
    def __init__(self, proj, path_building):
        self.proj= proj
        self.path_building= path_building 
        
    def df (self):
        proj = self.proj
        path_building = self.path_building
        config = configparser.ConfigParser()
        config.read('configurations.ini')
        if (proj == 'TLC' or proj == 'tlc' or proj == 'MMC' or proj == 'mmc'):
            path_proj = config['path']['tlc']
        if (proj == 'ORC' or proj == 'orc' or proj == 'PPK' or proj == 'ppk'): 
            path_proj = config['path']['orc']       
        list_folder = os.listdir(path_proj + r'/'+ path_building)    
        f = filter(lambda x: 'xlsx' in x, list_folder)
        list_folder = list(f)
        for i in range (len(list_folder)):
            if (r'~$' in list_folder[i]):
                list_folder.pop(i)
        df_folder = pd.Series(list_folder)
        df_folder = df_folder.str.split('.xlsx', expand=True)
        df_folder = df_folder.drop(1, axis = 1)
        df_folder
        df_folder_codes = df_folder[0].str.split('_', expand=True)
        df_folder_codes.rename(columns= {0: 'project', 
        1: 'stage',
        2: 'phase',
        3: 'building',
        4: 'company',
        5: 'type_file',
        6: 'discipline',
        7: 'discipline_rus',
        8: 'description'},
        inplace=True)
        xls_list = list(list_folder)
        listsht = list(list_folder)
        listdf = list(list_folder)
        df_mc = list(list_folder)
        df_mcmt = list(list_folder)
        df_m = list(list_folder)
        df_room = list(list_folder)
        for i in range(len(list_folder)):
            xls_list[i] = openpyxl.load_workbook(path_proj + r'/' + path_building + r'/' + list_folder[i])
            listsht[i] = xls_list[i].sheetnames
            listdf[i] = list(listsht[i])
        #Заполнение списка датафремов 
            for j in range(len(listsht[i])): 
                listdf[i][j] = pd.read_excel(path_proj + r'/' + path_building + r'/' + list_folder[i], sheet_name = listsht[i][j])
                if (listdf[i][j].size == 0):
                    listdf[i][j] = pd.DataFrame(columns=['Family and Type', 'Count'])  
            for j in range(len(listsht[i])): 
                listdf[i][j] = pd.read_excel(path_proj + r'/' + path_building + r'/' + list_folder[i], sheet_name = listsht[i][j])
                if (listdf[i][j].size == 0):
                    listdf[i][j] = pd.DataFrame(columns=['Family and Type', 'Count']) 
        #Заполнение пустых спецификаций
            for j in range(len(listsht[i])): 
                listdf[i][j] = pd.read_excel(path_proj + r'/' + path_building + r'/' + list_folder[i], sheet_name = listsht[i][j])
                if (listdf[i][j].size == 0 or 'Department' in listdf[i][j].columns):
                    listdf[i][j] = pd.DataFrame(columns=['Family and Type', 'Count']) 
        #Установка мульти категорий           
            df_mc[i] = pd.read_excel(path_proj + r'/' + path_building + r'/' + list_folder[i], sheet_name = 'VTBR_Multi-Category')
            df_mcmt[i] = pd.read_excel(path_proj + r'/' + path_building + r'/' + list_folder[i], sheet_name = 'VTBR_Multi-Category MT')
        #Объединение мульти и мульти материал
            df_m[i] = pd.concat([df_mcmt[i],df_mc[i]], join = "outer")  
        #Установка датафр помещений  
            if 'VTBR_Room' in listsht[i]:
                df_room[i] = pd.read_excel(path_proj + r'/' + path_building + r'/' + list_folder[i], sheet_name = 'VTBR_Room') 
        #Удаление мульти спец 
            for j in range(len(listsht[i])):         
                if ('Category' in listdf[i][j].columns):
                    listdf[i].append(listdf[i][j])
                    del listdf[i][j]
            for j in range(len(listsht[i])):           
                if ('Category' in listdf[i][j].columns):
                    listdf[i].append(listdf[i][j])
                    del listdf[i][j]
            del listdf[i][-2:]    
        #Добавление Category 
        for i in range(len(listdf)):
            for j in range(len(listdf[i])):          
                    listdf[i][j].insert(0, 'Building', df_folder_codes.building[i])
                    listdf[i][j].insert(1, 'Discipline', df_folder_codes.discipline[i])
                    listdf[i][j].insert(2, 'Category', (listdf[i][j].merge(df_m[i], on = 'Family and Type', how = 'left').Category))
                    listdf[i][j].insert(3, 'File', list_folder[i])            
        #Объединение спецификаций 
        df_cnc = list(listdf)
        for i in range(len(listdf)):
            df_cnc[i] = pd.concat([listdf[i][0]
                                   ], join = "outer")
            for j in range(1, len(listdf[i])):
                df_cnc[i] = pd.concat([df_cnc[i],
                                    listdf[i][j]
                                   ], join = "outer")

        df_cnc_building = pd.concat([df_cnc[0]
                                   ], join = "outer")
        for i in range(1, len(df_cnc)):
            df_cnc_building = pd.concat([df_cnc_building,
                                    df_cnc[i]
                                   ], join = "outer")
        return df_cnc_building 
    
    def files(self):
        proj = self.proj
        path_building = self.path_building
        config = configparser.ConfigParser()
        config.read('configurations.ini')
        if (proj == 'TLC' or proj == 'tlc'):
            path_proj = config['path']['tlc']
        if (proj == 'ORC' or proj == 'orc'): 
            path_proj = config['path']['orc'] 
        list_folder = os.listdir(path_proj + r'/'+ path_building)    
        f = filter(lambda x: 'xlsx' in x, list_folder)
        for i in range (len(list_folder)):
            if (r'~$' in list_folder[i]):
                list_folder.pop(i)
        list_folder = list(f)
        return list_folder 
    
    def path(self):
        proj = self.proj
        path_building = self.path_building
        config = configparser.ConfigParser()
        config.read('configurations.ini')
        if (proj == 'TLC' or proj == 'tlc'):
            path_proj = config['path']['tlc']
        if (proj == 'ORC' or proj == 'orc'): 
            path_proj = config['path']['orc'] 
        return path_proj + '/' + path_building
    
    def category(self):
        proj = self.proj
        path_building = self.path_building
        config = configparser.ConfigParser()
        config.read('configurations.ini')
        if (proj == 'TLC' or proj == 'tlc'):
            path_proj = config['path']['tlc']
        if (proj == 'ORC' or proj == 'orc'): 
            path_proj = config['path']['orc']
        list_folder = os.listdir(path_proj + r'/'+ path_building)    
        f = filter(lambda x: 'xlsx' in x, list_folder)
        for i in range (len(list_folder)):
            if (r'~$' in list_folder[i]):
                list_folder.pop(i)
        list_folder = list(f) 
        xls_list = list(list_folder)
        listsht = list(list_folder)
        for i in range(len(list_folder)):
            xls_list[i] = openpyxl.load_workbook(path_proj + r'/' + path_building + r'/' + list_folder[i])
            listsht[i] = xls_list[i].sheetnames            
        return listsht

    def check_family(self):
        proj = self.proj
        path_building = self.path_building
        config = configparser.ConfigParser()
        config.read('configurations.ini')
        if (proj == 'TLC' or proj == 'tlc' or proj == 'MMC' or proj == 'mmc'):
            path_proj = config['path']['tlc']
        if (proj == 'ORC' or proj == 'orc' or proj == 'PPK' or proj == 'ppk'): 
            path_proj = config['path']['orc']       
        list_folder = os.listdir(path_proj + r'/'+ path_building)    
        f = filter(lambda x: 'xlsx' in x, list_folder)
        list_folder = list(f)
        for i in range (len(list_folder)):
            if (r'~$' in list_folder[i]):
                list_folder.pop(i)
        df_folder = pd.Series(list_folder)
        df_folder = df_folder.str.split('.xlsx', expand=True)
        df_folder = df_folder.drop(1, axis = 1)
        df_folder
        df_folder_codes = df_folder[0].str.split('_', expand=True)
        df_folder_codes.rename(columns= {0: 'project', 
        1: 'stage',
        2: 'phase',
        3: 'building',
        4: 'company',
        5: 'type_file',
        6: 'discipline',
        7: 'discipline_rus',
        8: 'description'},
        inplace=True)
        xls_list = list(list_folder)
        listsht = list(list_folder)
        listdf = list(list_folder)
        for i in range(len(list_folder)):
            xls_list[i] = openpyxl.load_workbook(path_proj + r'/' + path_building + r'/' + list_folder[i])
            listsht[i] = xls_list[i].sheetnames
            listdf[i] = list(listsht[i])
        #Заполнение списка датафремов 
            for j in range(len(listsht[i])): 
                listdf[i][j] = pd.read_excel(path_proj + r'/' + path_building + r'/' + list_folder[i], sheet_name = listsht[i][j])
                if (listdf[i][j].size == 0):
                    listdf[i][j] = pd.DataFrame(columns=['Family and Type', 'Count'])  
            for j in range(len(listsht[i])): 
                listdf[i][j] = pd.read_excel(path_proj + r'/' + path_building + r'/' + list_folder[i], sheet_name = listsht[i][j])
                if (listdf[i][j].size == 0):
                    listdf[i][j] = pd.DataFrame(columns=['Family and Type', 'Count']) 
        #Заполнение пустых спецификаций
            for j in range(len(listsht[i])): 
                listdf[i][j] = pd.read_excel(path_proj + r'/' + path_building + r'/' + list_folder[i], sheet_name = listsht[i][j])
                if (listdf[i][j].size == 0 or 'Department' in listdf[i][j].columns):
                    listdf[i][j] = pd.DataFrame(columns=['Family and Type', 'Count'])  
    #Получение списка проверки        
        true_lst = copy.deepcopy(listdf)
        for i in range(len(listdf)):
            for j in range(len(listdf[i])):
                true_lst[i][j] = ('Family and Type' in listdf[i][j].columns)
        check_list = list(true_lst)
        for i in range(len(listdf)):
            check_list[i] = all(true_lst[i])  
        return check_list

    def check_multi(self):
        #Проверка мульти категорий
        proj = self.proj
        path_building = self.path_building
        config = configparser.ConfigParser()
        config.read('configurations.ini')
        if (proj == 'TLC' or proj == 'tlc' or proj == 'MMC' or proj == 'mmc'):
            path_proj = config['path']['tlc']
        if (proj == 'ORC' or proj == 'orc' or proj == 'PPK' or proj == 'ppk'): 
            path_proj = config['path']['orc']       
        list_folder = os.listdir(path_proj + r'/'+ path_building)    
        f = filter(lambda x: 'xlsx' in x, list_folder)
        list_folder = list(f)
        for i in range (len(list_folder)):
            if (r'~$' in list_folder[i]):
                list_folder.pop(i)
        xls_list = list(list_folder)
        listsht = list(list_folder)
        for i in range(len(list_folder)):
            xls_list[i] = openpyxl.load_workbook(path_proj + r'/' + path_building + r'/' + list_folder[i])
            listsht[i] = xls_list[i].sheetnames        
        list_true_mt = list(list_folder)
        for i in range(len(list_folder)):
            list_true_mt[i] = 'VTBR_Multi-Category MT' in listsht[i]
        return list_true_mt       

